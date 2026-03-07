import csv
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from quest_app.models import Exercise,ExerciseCategory

class Command(BaseCommand):
    help = "Import exercises and link a single existing category (M2M)"

    def handle(self, *args, **kwargs):
        # with open('quest_app/exercises.xlsx', newline='', encoding='utf-8') as file:
        #     reader = csv.DictReader(file)

        #     for row in reader:
        #         exercise, created = Exercise.objects.get_or_create(
        #             exercise_name=row['exercise_name'],
        #             target_zones_major=row['target_zones_major'],
        #             target_zones_minor=row['target_zones_minor'],
        #             equipment=row['equipment']
        #         )

        #         # Optional: Clear previous categories before adding new one
        #         exercise.category.clear()

        #         try:
        #             category = ExerciseCategory.objects.get(id=row['category'])
        #             exercise.category.add(category)
                    
        #         except ExerciseCategory.DoesNotExist:
        #             self.stdout.write(self.style.ERROR(f"❌ Category ID {row['category']} not found."))

        #         exercise.save()
        #         action = "🟢 Created" if created else "🟡 Updated"
        #         self.stdout.write(f"{action} Exercise: {exercise.exercise_name}")


        workbook = load_workbook('quest_app/exercises.xlsx')
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            data = dict(zip(headers, row))

            exercise, created = Exercise.objects.get_or_create(
                exercise_name=data['Exercise Name'],
                target_zones_major=data['Target Zone (Major)'],
                target_zones_minor=data['Target Zone (Minor)'],
                equipment=data['Equipment']
            )

            # Clear existing categories
            exercise.category.clear()

            try:
                category = ExerciseCategory.objects.get(id=data['Category ID'])
                exercise.category.add(category)

            except ExerciseCategory.DoesNotExist:
                self.stdout.write(
                    self.style.ERROR(f"❌ Category ID {data['Category ID']} not found.")
                )

            exercise.save()

            action = "🟢 Created" if created else "🟡 Updated"
            self.stdout.write(f"{action} Exercise: {exercise.exercise_name}")
